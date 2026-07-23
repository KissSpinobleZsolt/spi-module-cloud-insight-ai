"""
CloudInsight AI — plugin backend

Provides bot-config persistence and file ingestion for the CloudInsight AI module.
The core backend proxies all requests here via /api/plugin/cloudInsightAI/...
Authentication: the Authorization header is forwarded verbatim from the proxy —
every protected route validates it with verify_token() before touching data.

Environment:
  JWT_SECRET_KEY   — shared with the core backend (same value in docker-compose)
  BOT_CONFIGS_FILE — path to the JSON file used as a lightweight bot config store
  CORE_API_URL     — base URL of the core backend for log writes (default: http://backend:8000)
"""

import csv
import io
import json
import os
from typing import Any, Optional

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from jose import JWTError, jwt
from pydantic import BaseModel, Field

import spin_logger

JWT_SECRET = os.getenv("JWT_SECRET_KEY", "change-me-in-production")
ALGORITHM = "HS256"
BOT_CONFIGS_FILE = os.getenv("BOT_CONFIGS_FILE", "/app/data/bot_configs.json")  # lightweight JSON config store

app = FastAPI(title="CloudInsight AI Backend", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def verify_token(authorization: str | None = Header(default=None)):
    """Validate the JWT forwarded by the core proxy."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1]
    try:
        jwt.decode(token, JWT_SECRET, algorithms=[ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")


def _load_configs() -> dict:
    """Read bot configs from disk; return empty dict if file is missing or corrupt."""
    try:
        with open(BOT_CONFIGS_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_configs(data: dict) -> None:
    """Persist bot configs to disk, creating parent dirs as needed."""
    os.makedirs(os.path.dirname(BOT_CONFIGS_FILE), exist_ok=True)
    with open(BOT_CONFIGS_FILE, "w") as f:
        json.dump(data, f)


class ConfigUpdatePayload(BaseModel):
    config: dict[str, Any]  # arbitrary key-value config from the frontend form


class IngestionJobConfig(BaseModel):
    """Config schema mirroring the CloudInsight Monitor bot's manifest config_schema."""
    error_threshold_pct: float = Field(default=5.0, ge=0, le=100)  # max % of bad rows before flagging
    stale_hours: int = Field(default=24, ge=1)  # hours before a source is considered stale
    min_row_count: int = Field(default=1, ge=0)  # reject ingestion if record count is below this
    notify_on_quality: bool = True  # emit a quality alert when threshold is breached
    scan_interval_minutes: int = Field(default=60, ge=5, le=120)  # scheduler cadence


class IngestionResult(BaseModel):
    message: str  # human-readable outcome used directly in the frontend toast
    record_count: Optional[int] = None  # number of parsed records (None for XLSX without openpyxl)
    file_size_bytes: int  # raw byte length of the uploaded file
    status: str  # "success" | "error"


def _count_records(content: bytes, ext: str) -> int:
    """Parse the file content and return a record count appropriate to the format."""
    if ext == "csv":
        reader = csv.reader(io.StringIO(content.decode("utf-8", errors="replace")))
        rows = list(reader)
        return max(0, len(rows) - 1)  # subtract header row

    if ext == "json":
        parsed = json.loads(content)
        if isinstance(parsed, list):
            return len(parsed)  # array of objects → one record each
        if isinstance(parsed, dict):
            return len(parsed)  # top-level keys treated as record identifiers
        return 1  # scalar or other

    if ext == "xlsx":
        from openpyxl import load_workbook  # imported here to surface ImportError clearly
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        count = max(0, (ws.max_row or 1) - 1)  # subtract header row
        wb.close()
        return count

    raise ValueError(f"Unsupported extension: {ext}")


@app.get("/health")
def health():
    """Liveness probe used by docker-compose healthcheck."""
    return {"status": "ok", "service": "cloud-insight-ai-backend"}


@app.get("/bots/{bot_uuid}/config")
async def get_bot_config(
    bot_uuid: str,
    authorization: str | None = Header(default=None),
    _: None = Depends(verify_token),
):
    """Return stored config for a bot; returns empty config if not yet saved."""
    spin_logger.log_module(authorization or "", "module.activated", {"action": "bot_config_read", "bot_uuid": bot_uuid})  # log module activation when the frontend loads bot config
    configs = _load_configs()
    return {"config": configs.get(bot_uuid, {}), "teams": []}  # teams not used by CloudInsight AI bots


@app.put("/bots/{bot_uuid}/config")
async def update_bot_config(
    bot_uuid: str,
    payload: ConfigUpdatePayload,
    authorization: str | None = Header(default=None),
    _: None = Depends(verify_token),
):
    """Persist the full config for a bot (replaces previous value)."""
    configs = _load_configs()
    configs[bot_uuid] = payload.config  # full replacement, matching the core pattern
    _save_configs(configs)
    spin_logger.log_bot(  # record config change against the bot's own log
        authorization or "",
        bot_uuid,
        "bot.config.updated",
        payload.config,
        message="Bot configuration updated via CloudInsight AI",
    )
    return {"config": configs[bot_uuid]}


@app.post("/upload", response_model=IngestionResult)
async def upload_file(
    file: UploadFile = File(...),
    bot_identifier: str = Form(default="CloudInsight Monitor"),  # audit label; not yet resolved to a UUID
    authorization: str | None = Header(default=None),
    _: None = Depends(verify_token),
):
    """
    Accept a CSV, XLSX, or JSON file, parse it using the CloudInsight Monitor
    bot's IngestionJobConfig, and return a structured ingestion result.
    """
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in {"csv", "xlsx", "json"}:
        raise HTTPException(
            status_code=422,
            detail=f"data was errored: unsupported file type .{ext} — use CSV, XLSX, or JSON",
        )

    spin_logger.log_module(  # log before reading so the event is recorded even if parsing fails
        authorization or "",
        "ingest.started",
        {"filename": file.filename, "ext": ext, "bot_identifier": bot_identifier},
    )

    content = await file.read()
    file_size = len(content)

    config = IngestionJobConfig()  # use default thresholds; a future version can load per bot_uuid

    try:
        record_count = _count_records(content, ext)
    except Exception as exc:
        result = IngestionResult(
            message=f"data was errored: {exc}",
            file_size_bytes=file_size,
            status="error",
        )
        spin_logger.log_module(
            authorization or "",
            "ingest.failed",
            {"filename": file.filename, "error": str(exc), "file_size_bytes": file_size},
        )
        return result

    # Reject ingestion if the record count is below the configured minimum threshold
    if record_count < config.min_row_count:
        result = IngestionResult(
            message=(
                f"data was errored: file has {record_count:,} records, "
                f"below the minimum threshold of {config.min_row_count}"
            ),
            record_count=record_count,
            file_size_bytes=file_size,
            status="error",
        )
        spin_logger.log_module(
            authorization or "",
            "ingest.rejected",
            {"filename": file.filename, "record_count": record_count, "min_row_count": config.min_row_count},
        )
        return result

    spin_logger.log_module(  # log successful ingestion with counts for observability
        authorization or "",
        "ingest.completed",
        {"filename": file.filename, "record_count": record_count, "file_size_bytes": file_size},
    )

    return IngestionResult(
        message=f"data was consumed and had {record_count:,} records ({file_size:,} bytes)",
        record_count=record_count,
        file_size_bytes=file_size,
        status="success",
    )
